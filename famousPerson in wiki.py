# pip install wikipedia

import openpyxl
import requests
import time
import wikipedia
from datetime import datetime

# 设置语言为中文
wikipedia.set_lang("zh")


# 定义获取词条状态的函数
def get_wiki_status(keywords):
    count = 0
    count_threshold = 100
    url_prefix = "https://zh.wikipedia.org/wiki/"
    output = []
    for keyword in keywords:
        count += 1
        if count % count_threshold == 0:
            print(str(count) + " records finished...")
        url = url_prefix + keyword
        response = requests.head(url)
        time.sleep(0.1)
        if response.status_code == 200:
            output.append([keyword, "1"])
        else:
            output.append([keyword, "0"])
    return output


# 读取关键词文件
with open("input.txt", "r", encoding="utf-8-sig") as file:
    keywords = file.read().splitlines()

# 创建Excel文件
wb = openpyxl.Workbook()
sheet = wb.active

# 写入表头
sheet["A1"] = "id"
sheet["B1"] = "name"
sheet["C1"] = "exist_status"
sheet["D1"] = "content"
sheet["E1"] = "brief_content"
sheet["F1"] = "timestamp"  # 新增的列

# 遍历关键词列表
row = 2  # 行号
for i, keyword in enumerate(keywords):
    try:
        # 使用get_wiki_status函数判断关键词是否存在于Wikipedia词条
        status = get_wiki_status([keyword])[0][1]

        if status == "1":
            # 若存在，则获取词条内容
            page = wikipedia.page(keyword)
            content = page.content
            brief_content = page.content.split("\n")[0]

            # 写入结果到Excel
            sheet.cell(row=row, column=1).value = i + 1
            sheet.cell(row=row, column=2).value = keyword
            sheet.cell(row=row, column=3).value = status
            sheet.cell(row=row, column=4).value = content
            sheet.cell(row=row, column=5).value = brief_content
            sheet.cell(row=row, column=6).value = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )  # 写入当前时间
            row += 1
        else:
            # 若不存在，则设置内容为空
            sheet.cell(row=row, column=1).value = i + 1
            sheet.cell(row=row, column=2).value = keyword
            sheet.cell(row=row, column=3).value = status
            sheet.cell(row=row, column=6).value = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )  # 写入当前时间
            row += 1

    except wikipedia.exceptions.DisambiguationError as e:
        # 处理歧义词条，将每个选项作为单独的行输出
        options = e.options
        for j, option in enumerate(options):
            try:
                # 若存在，则获取词条内容
                page = wikipedia.page(option)
                content = page.content
                brief_content = page.content.split("\n")[0]
            except (
                wikipedia.exceptions.DisambiguationError,
                wikipedia.exceptions.PageError,
            ):
                continue

            # 写入结果到Excel
            sheet.cell(row=row, column=1).value = i + 1
            sheet.cell(row=row, column=2).value = keyword + " - " + str(j + 1)
            sheet.cell(row=row, column=3).value = status
            sheet.cell(row=row, column=4).value = content
            sheet.cell(row=row, column=5).value = brief_content
            sheet.cell(row=row, column=6).value = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )  # 写入当前时间
            row += 1


# 保存Excel文件
wb.save("wiki_result.xlsx")
print("结果已保存为wiki_result.xlsx")
